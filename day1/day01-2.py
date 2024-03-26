import os
import openpyxl
path=r"D:\code\pythoncode\pythonexcel"
os.chdir(path)
wb=openpyxl.load_workbook("test001.xlsx")
sheet=wb.active # get active sheet
print(sheet.max_row)
print(sheet.max_column)
cell1=sheet.cell(row=3,column=3) # print cell value
print(cell1.value)


